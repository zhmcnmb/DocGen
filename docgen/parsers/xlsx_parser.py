from openpyxl import load_workbook
from pathlib import Path


def parse_xlsx(path: str) -> str:
    filename = Path(path).name
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cells):
                parts.append(" | ".join(cells))

    wb.close()
    content = "\n".join(parts)
    return f"[文件: {filename}]\n{content}"
